"""
Excel parser handling:
  - Tier 1: Simple tabular (pandas)
  - Tier 2: Merged/multi-level headers (openpyxl + merge-fill)
  - Tier 3: Bilingual mirrored tables (language detection + column pairing)

Returns structured data for storage in StructuredTable,
plus a natural language description for embedding.
"""
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional
import math
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from loguru import logger

from pipeline.ingestion.processors.arabic_normalizer import detect_language


@dataclass
class ExtractedTable:
    sheet_name: str
    table_index: int
    columns_ar: list[dict]           # [{"name": "المنطقة", "type": "text"}]
    columns_en: list[dict]           # [{"name": "Region", "type": "text"}]
    data: list[dict]                 # list of row dicts
    row_count: int
    natural_language_description: str = ""


def parse_excel(file_path: Path) -> list[ExtractedTable]:
    """Parse all sheets of an Excel file and return structured tables."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    tables: list[ExtractedTable] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            sheet_tables = _parse_sheet(ws, sheet_name)
            tables.extend(sheet_tables)
        except Exception as e:
            logger.warning(f"Failed to parse sheet '{sheet_name}' in {file_path}: {e}")

    return tables


def _parse_sheet(ws, sheet_name: str) -> list["ExtractedTable"]:
    # Step 1: Fill merged cells forward
    _fill_merged_cells(ws)

    # Step 2: Convert to raw dataframe
    data = list(ws.values)
    if not data or len(data) < 2:
        return []

    df_raw = pd.DataFrame(data)
    df_raw = df_raw.dropna(how="all").dropna(axis=1, how="all")
    if df_raw.empty:
        return []

    # Step 3: Detect header row(s) and complexity
    header_rows, data_start = _detect_header_rows(df_raw)
    complexity = _measure_complexity(df_raw, header_rows)

    if complexity == "bilingual":
        return _parse_bilingual_sheet(df_raw, header_rows, data_start, sheet_name)
    else:
        return _parse_standard_sheet(df_raw, header_rows, data_start, sheet_name, complexity)


def _fill_merged_cells(ws):
    """
    openpyxl leaves merged cells with None except for the top-left cell.
    Fill all cells in each merged range with the top-left value.
    Non-top-left cells are MergedCell objects (read-only), so unmerge first.
    """
    for merged_range in list(ws.merged_cells.ranges):
        top_left_value = ws.cell(merged_range.min_row, merged_range.min_col).value
        # Unmerge so non-top-left cells become writable regular cells
        ws.unmerge_cells(str(merged_range))
        for row in ws.iter_rows(
            min_row=merged_range.min_row,
            max_row=merged_range.max_row,
            min_col=merged_range.min_col,
            max_col=merged_range.max_col,
        ):
            for cell in row:
                cell.value = top_left_value


def _detect_header_rows(df: pd.DataFrame) -> tuple[list[int], int]:
    """
    Detect which rows are headers (short text values, no numbers, high unique ratio).
    Returns (list of header row indices, index of first data row).
    """
    header_rows = []
    for i, row in df.iterrows():
        values = [v for v in row if v is not None and str(v).strip()]
        if not values:
            continue
        numeric_count = sum(1 for v in values if _is_numeric(v))
        text_count = len(values) - numeric_count
        if text_count / len(values) > 0.7 and i < 5:
            header_rows.append(i)
        else:
            break

    data_start = max(header_rows) + 1 if header_rows else 0
    return header_rows, data_start


def _measure_complexity(df: pd.DataFrame, header_rows: list[int]) -> str:
    """
    Determine if this is a standard, multi-level, or bilingual table.
    Returns: 'simple' | 'multilevel' | 'bilingual'
    """
    if not header_rows:
        return "simple"

    if len(header_rows) > 1:
        return "multilevel"

    # Bilingual detection: check if columns alternate Arabic/English
    header_row = df.iloc[header_rows[0]] if header_rows else df.iloc[0]
    col_langs = [detect_language(str(v)) for v in header_row if v is not None]
    ar_count = col_langs.count("ar")
    en_count = col_langs.count("en")

    if ar_count > 0 and en_count > 0:
        ar_ratio = ar_count / (ar_count + en_count)
        if 0.3 < ar_ratio < 0.7:   # roughly balanced
            return "bilingual"

    return "simple"


def _parse_standard_sheet(
    df: pd.DataFrame,
    header_rows: list[int],
    data_start: int,
    sheet_name: str,
    complexity: str,
) -> list["ExtractedTable"]:
    if not header_rows:
        columns = [str(i) for i in range(len(df.columns))]
    elif complexity == "multilevel":
        columns = _flatten_multilevel_headers(df, header_rows)
    else:
        columns = [str(v) for v in df.iloc[header_rows[0]]]

    data_df = df.iloc[data_start:].copy()
    data_df.columns = columns[:len(data_df.columns)]
    data_df = data_df.dropna(how="all")

    # Classify each column as AR or EN
    cols_ar = [{"name": c, "type": _infer_col_type(data_df[c])} for c in columns if detect_language(c) in ("ar", "unknown")]
    cols_en = [{"name": c, "type": _infer_col_type(data_df[c])} for c in columns if detect_language(c) == "en"]

    rows = data_df.to_dict(orient="records")
    rows = [{k: _serialize_value(v) for k, v in row.items()} for row in rows]

    return [ExtractedTable(
        sheet_name=sheet_name,
        table_index=0,
        columns_ar=cols_ar,
        columns_en=cols_en,
        data=rows,
        row_count=len(rows),
        natural_language_description=_generate_description(sheet_name, columns, rows[:3]),
    )]


def _parse_bilingual_sheet(
    df: pd.DataFrame,
    header_rows: list[int],
    data_start: int,
    sheet_name: str,
) -> list["ExtractedTable"]:
    """
    Handle bilingual mirrored tables where Arabic and English columns are paired.
    Pairs AR and EN columns, creates a unified bilingual schema.
    """
    header_row = df.iloc[header_rows[0]] if header_rows else df.iloc[0]
    col_values = list(header_row)

    ar_cols = [(i, str(v)) for i, v in enumerate(col_values) if detect_language(str(v)) == "ar"]
    en_cols = [(i, str(v)) for i, v in enumerate(col_values) if detect_language(str(v)) == "en"]
    num_cols = [(i, str(v)) for i, v in enumerate(col_values) if _is_numeric(v) or v is None]

    data_df = df.iloc[data_start:].copy()

    # Build unified rows with _ar / _en suffixed column names
    unified_rows = []
    for _, row in data_df.iterrows():
        unified_row = {}
        for idx, ar_name in ar_cols:
            if idx < len(row):
                unified_row[f"{ar_name}_ar"] = _serialize_value(row.iloc[idx])
        for idx, en_name in en_cols:
            if idx < len(row):
                unified_row[f"{en_name}_en"] = _serialize_value(row.iloc[idx])
        for idx, num_name in num_cols:
            if idx < len(row) and row.iloc[idx] is not None:
                unified_row[num_name or f"col_{idx}"] = _serialize_value(row.iloc[idx])
        if any(v is not None for v in unified_row.values()):
            unified_rows.append(unified_row)

    cols_ar = [{"name": f"{n}_ar", "type": "text"} for _, n in ar_cols]
    cols_en = [{"name": f"{n}_en", "type": "text"} for _, n in en_cols]

    all_col_names = list(unified_rows[0].keys()) if unified_rows else []
    return [ExtractedTable(
        sheet_name=sheet_name,
        table_index=0,
        columns_ar=cols_ar,
        columns_en=cols_en,
        data=unified_rows,
        row_count=len(unified_rows),
        natural_language_description=_generate_description(sheet_name, all_col_names, unified_rows[:3]),
    )]


def _flatten_multilevel_headers(df: pd.DataFrame, header_rows: list[int]) -> list[str]:
    """Flatten multi-level headers into compound strings like 'Hospitals | Public'."""
    header_data = df.iloc[header_rows]
    flat = []
    for col_idx in range(len(df.columns)):
        parts = []
        for row_idx in header_rows:
            val = df.iloc[row_idx, col_idx]
            if val is not None and str(val).strip():
                s = str(val).strip()
                if not parts or s != parts[-1]:   # deduplicate repeated merged values
                    parts.append(s)
        flat.append(" | ".join(parts) if parts else f"col_{col_idx}")
    return flat


def _infer_col_type(series: pd.Series) -> str:
    non_null = series.dropna()
    if non_null.empty:
        return "text"
    numeric_count = sum(1 for v in non_null if _is_numeric(v))
    return "number" if numeric_count / len(non_null) > 0.7 else "text"


def _is_numeric(value) -> bool:
    if value is None:
        return False
    try:
        float(str(value).replace(",", "").replace("،", ""))
        return True
    except (ValueError, TypeError):
        return False


def _serialize_value(value):
    if value is None:
        return None
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return value
    if isinstance(value, int):
        return value
    return str(value).strip()


def _generate_description(sheet_name: str, columns: list[str], sample_rows: list[dict]) -> str:
    """Generate a natural language description for the table, used as the embedded text."""
    col_str = ", ".join(c for c in columns[:12] if c)
    # Format sample rows as readable key=value pairs rather than raw dict repr
    sample_parts = []
    for row in sample_rows[:3]:
        pairs = "; ".join(f"{k}: {v}" for k, v in row.items() if v is not None and str(v).strip())
        if pairs:
            sample_parts.append(pairs)
    sample_str = " | ".join(sample_parts)
    return (
        f"{sheet_name}. "
        f"Columns: {col_str}. "
        f"Sample rows: {sample_str}"
    )
